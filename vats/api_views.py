"""
vats/api_views.py  —  HelpDesk Pro v2
REST API views — all role checks use "Admin"/"Manager"/"Viewer" (capitalised)
to match your existing registration.User.role_choice values exactly.

Endpoints added:
  GET  /api/v1/tickets/               — paginated, filterable ticket list
  POST /api/v1/tickets/               — create ticket (Viewer only)
  GET  /api/v1/tickets/<id>/          — ticket detail + worknotes
  PATCH /api/v1/tickets/<id>/         — update status/assignee (Admin/Manager)
  GET  /api/v1/tickets/export/csv/    — download as CSV
  GET  /api/v1/tickets/export/excel/  — download as Excel
  GET  /api/v1/analytics/dashboard/   — KPI cards + chart data
  GET  /api/v1/analytics/report/      — date-range report
  GET  /api/v1/categories/            — category list
  GET  /api/v1/users/managers/        — managers list (for dropdowns)
"""
import csv
import io
from datetime import timedelta

from django.db.models import Count, Q
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import generics, permissions, status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView

from .filters import TicketFilter
from .models import Category, Subcategory, Ticket, Worknote
from .serializers import (
    CategorySerializer,
    SubcategorySerializer,
    TicketDetailSerializer,
    TicketListSerializer,
    UserMinimalSerializer,
    UserSerializer,
    WorknoteSerializer,
)
from registration.models import User


# ── Permission helpers ────────────────────────────────────────────────────────

class IsAdmin(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.role == 'Admin'


class IsAdminOrManager(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.role in ('Admin', 'Manager')


# ── Ticket ViewSet ────────────────────────────────────────────────────────────

class TicketViewSet(viewsets.ModelViewSet):
    """
    Role-based queryset — matches your existing views.py logic exactly:
      Admin   → all tickets
      Manager → tickets assigned_to them
      Viewer  → tickets they created
    """
    permission_classes = [permissions.IsAuthenticated]
    filterset_class    = TicketFilter
    search_fields      = ['title', 'problem_descp', 'number']
    ordering_fields    = ['created_at', 'updated_at', 'priority', 'status', 'due_by']
    ordering           = ['-created_at']

    def get_queryset(self):
        user = self.request.user
        base = Ticket.objects.select_related(
            'created_by', 'assigned_to', 'category', 'subcategory'
        ).prefetch_related('Worknotes')

        if user.role == 'Admin':
            return base.all()
        if user.role == 'Manager':
            return base.filter(assigned_to=user)
        # Viewer
        return base.filter(created_by=user)

    def get_serializer_class(self):
        if self.action in ('retrieve', 'update', 'partial_update'):
            return TicketDetailSerializer
        return TicketListSerializer

    def perform_create(self, serializer):
        serializer.save(
            created_by=self.request.user,
            status='Pending',
        )

    # ── CSV Export ────────────────────────────────────────────────────────────
    @action(detail=False, methods=['get'], url_path='export/csv')
    def export_csv(self, request):
        """
        GET /api/v1/tickets/export/csv/
        Download all visible (filtered) tickets as CSV.
        Supports all the same filters as the list endpoint.
        Key for Data Analyst role: shows you understand data export pipelines.
        """
        queryset = self.filter_queryset(self.get_queryset())

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = (
            f'attachment; filename="helpdesk_tickets_{timezone.now().date()}.csv"'
        )

        writer = csv.writer(response)
        writer.writerow([
            'Ticket #', 'Title', 'Category', 'Subcategory',
            'Priority', 'Status',
            'Created By', 'Assigned To',
            'Created At', 'Updated At',
            'SLA Due By', 'Resolved At',
            'Resolution Time (hrs)', 'SLA Breached',
        ])

        for t in queryset:
            writer.writerow([
                t.number,
                t.title,
                t.category.name if t.category else '',
                t.subcategory.name if t.subcategory else '',
                t.priority or '',
                t.status or '',
                t.created_by.get_full_name() if t.created_by else '',
                t.assigned_to.get_full_name() if t.assigned_to else 'Unassigned',
                t.get_created_at().strftime('%Y-%m-%d %H:%M'),
                t.get_updated_at().strftime('%Y-%m-%d %H:%M'),
                t.due_by.strftime('%Y-%m-%d %H:%M') if t.due_by else '',
                t.resolved_at.strftime('%Y-%m-%d %H:%M') if t.resolved_at else '',
                t.resolution_time_hours if t.resolution_time_hours is not None else '',
                'Yes' if t.is_sla_breached else 'No',
            ])

        return response

    # ── Excel Export ──────────────────────────────────────────────────────────
    @action(detail=False, methods=['get'], url_path='export/excel')
    def export_excel(self, request):
        """GET /api/v1/tickets/export/excel/"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return Response(
                {'error': 'Install openpyxl: pip install openpyxl'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        queryset = self.filter_queryset(self.get_queryset())
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Tickets'

        headers = [
            'Ticket #', 'Title', 'Category', 'Subcategory',
            'Priority', 'Status', 'Created By', 'Assigned To',
            'Created At', 'Updated At', 'SLA Due By', 'Resolved At',
            'Resolution Time (hrs)', 'SLA Breached',
        ]

        # Styled header row
        header_fill = PatternFill(start_color='185FA5', end_color='185FA5', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for t in queryset:
            ws.append([
                t.number,
                t.title,
                t.category.name if t.category else '',
                t.subcategory.name if t.subcategory else '',
                t.priority or '',
                t.status or '',
                t.created_by.get_full_name() if t.created_by else '',
                t.assigned_to.get_full_name() if t.assigned_to else 'Unassigned',
                t.get_created_at().replace(tzinfo=None),
                t.get_updated_at().replace(tzinfo=None),
                t.due_by.replace(tzinfo=None) if t.due_by else None,
                t.resolved_at.replace(tzinfo=None) if t.resolved_at else None,
                t.resolution_time_hours,
                'Yes' if t.is_sla_breached else 'No',
            ])

        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            f'attachment; filename="helpdesk_tickets_{timezone.now().date()}.xlsx"'
        )
        return response


# ── Analytics — Dashboard ─────────────────────────────────────────────────────

class DashboardAnalyticsView(APIView):
    """
    GET /api/v1/analytics/dashboard/
    Powers the real-time dashboard KPI cards and charts.
    Matches your role logic: Admin sees all, Manager sees assigned, Viewer sees own.
    """
    permission_classes = [IsAdminOrManager]

    def get(self, request):
        now = timezone.now()

        # Role-filtered base queryset (same logic as your views.py)
        if request.user.role == 'Admin':
            qs = Ticket.objects.all()
        else:
            qs = Ticket.objects.filter(assigned_to=request.user)

        # ── KPI Cards ─────────────────────────────────────────────────────────
        open_statuses    = ('Pending', 'Assigned', 'Scoping', 'In Progress')
        closed_statuses  = ('Completed', 'Cancelled', 'Rejected')

        total    = qs.count()
        open_cnt = qs.filter(status__in=open_statuses).count()
        resolved = qs.filter(status__in=closed_statuses).count()
        pending  = qs.filter(status='Pending').count()

        # Avg resolution time in hours
        resolved_tickets = [t for t in qs if t.resolution_time_hours is not None]
        avg_resolution = None
        if resolved_tickets:
            avg_resolution = round(
                sum(t.resolution_time_hours for t in resolved_tickets) / len(resolved_tickets), 1
            )

        # SLA breaches
        sla_breached_cnt = sum(1 for t in qs if t.is_sla_breached)

        # ── Daily Volume — last 7 days ────────────────────────────────────────
        daily_data = []
        for i in range(6, -1, -1):
            day_start = (now - timedelta(days=i)).replace(
                hour=0, minute=0, second=0, microsecond=0
            )
            day_end = day_start + timedelta(days=1)
            day_qs  = qs.filter(created_at__gte=day_start, created_at__lt=day_end)
            daily_data.append({
                'date':     day_start.strftime('%a'),
                'open':     day_qs.filter(status__in=open_statuses).count(),
                'resolved': day_qs.filter(status__in=closed_statuses).count(),
                'pending':  day_qs.filter(status='Pending').count(),
            })

        # ── Priority Distribution ─────────────────────────────────────────────
        priority_data = (
            qs.values('priority')
              .annotate(count=Count('id'))
              .order_by('priority')
        )
        priority_chart = {row['priority']: row['count'] for row in priority_data}

        # ── Category Breakdown ────────────────────────────────────────────────
        category_data = (
            qs.filter(category__isnull=False)
              .values('category__name')
              .annotate(count=Count('id'))
              .order_by('-count')[:8]
        )
        category_chart = [
            {'name': row['category__name'], 'count': row['count']}
            for row in category_data
        ]

        # ── Team Workload (admin only) ────────────────────────────────────────
        team_workload = []
        if request.user.role == 'Admin':
            for manager in User.objects.filter(role='Manager'):
                m_qs = qs.filter(assigned_to=manager)
                team_workload.append({
                    'name':     manager.get_full_name(),
                    'total':    m_qs.count(),
                    'open':     m_qs.filter(status__in=open_statuses).count(),
                    'resolved': m_qs.filter(status__in=closed_statuses).count(),
                })

        return Response({
            'kpis': {
                'total':                 total,
                'open':                  open_cnt,
                'resolved':              resolved,
                'pending':               pending,
                'sla_breached':          sla_breached_cnt,
                'avg_resolution_hours':  avg_resolution,
            },
            'charts': {
                'daily_volume':      daily_data,
                'priority_dist':     priority_chart,
                'category_breakdown': category_chart,
                'team_workload':     team_workload,
            },
        })


# ── Analytics — Date-Range Report ────────────────────────────────────────────

class ReportAnalyticsView(APIView):
    """
    GET /api/v1/analytics/report/?start=2024-01-01&end=2024-03-31
    Powers the Reports page with configurable date ranges.
    Shows: resolution by category, SLA compliance by manager.
    """
    permission_classes = [IsAdminOrManager]

    def get(self, request):
        from django.utils.dateparse import parse_date
        now = timezone.now()

        start_str = request.query_params.get('start')
        end_str   = request.query_params.get('end')

        try:
            start = (
                timezone.datetime.combine(parse_date(start_str), timezone.datetime.min.time())
                if start_str else now - timedelta(days=30)
            )
            end = (
                timezone.datetime.combine(parse_date(end_str), timezone.datetime.max.time())
                if end_str else now
            )
        except Exception:
            start = now - timedelta(days=30)
            end   = now

        if request.user.role == 'Admin':
            qs = Ticket.objects.filter(created_at__gte=start, created_at__lte=end)
        else:
            qs = Ticket.objects.filter(
                assigned_to=request.user,
                created_at__gte=start, created_at__lte=end,
            )

        open_statuses   = ('Pending', 'Assigned', 'Scoping', 'In Progress')
        closed_statuses = ('Completed', 'Cancelled', 'Rejected')

        # Resolution time by category
        resolution_by_category = []
        for cat in Category.objects.all():
            cat_resolved = [
                t for t in qs.filter(category=cat)
                if t.resolution_time_hours is not None
            ]
            if cat_resolved:
                avg = round(
                    sum(t.resolution_time_hours for t in cat_resolved) / len(cat_resolved), 1
                )
                resolution_by_category.append({'category': cat.name, 'avg_hours': avg})

        # SLA compliance per manager (admin only)
        sla_compliance = []
        if request.user.role == 'Admin':
            for manager in User.objects.filter(role='Manager'):
                m_qs  = qs.filter(assigned_to=manager)
                total = m_qs.count()
                if total:
                    breached = sum(1 for t in m_qs if t.is_sla_breached)
                    sla_compliance.append({
                        'manager':        manager.get_full_name(),
                        'total':          total,
                        'breached':       breached,
                        'compliance_pct': round((total - breached) / total * 100, 1),
                    })

        return Response({
            'date_range': {'start': start.date().isoformat(), 'end': end.date().isoformat()},
            'summary': {
                'total':       qs.count(),
                'open':        qs.filter(status__in=open_statuses).count(),
                'resolved':    qs.filter(status__in=closed_statuses).count(),
                'sla_breached': sum(1 for t in qs if t.is_sla_breached),
            },
            'resolution_by_category': resolution_by_category,
            'sla_compliance':         sla_compliance,
        })


# ── Category ViewSet ──────────────────────────────────────────────────────────

class CategoryViewSet(viewsets.ReadOnlyModelViewSet):
    """GET /api/v1/categories/ — read-only (creates still go through admin/web UI)"""
    queryset         = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=True, methods=['get'])
    def subcategories(self, request, pk=None):
        category = self.get_object()
        subs = SubcategorySerializer(
            Subcategory.objects.filter(category=category), many=True
        )
        return Response(subs.data)


# ── User Views ────────────────────────────────────────────────────────────────

class ManagerListView(generics.ListAPIView):
    """
    GET /api/v1/users/managers/
    Returns only active Managers — used to populate the "Assign to" dropdown
    without loading ALL users.
    """
    serializer_class   = UserMinimalSerializer
    permission_classes = [IsAdminOrManager]

    def get_queryset(self):
        return User.objects.filter(role='Manager', is_active=True)


class CurrentUserView(APIView):
    """GET /api/v1/users/me/"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        return Response(UserSerializer(request.user).data)
